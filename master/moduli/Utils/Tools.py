import openpyxl
import os

import json
import time
from tkinter import Tk, Button
from tkinter.filedialog import askopenfilename


def json_to_excel_pg():
    """
    Utile per trasformare un json del tipo {"indice":"a","val1" :2...]} in una o piu colonne di excel
    usato principalmente per la lista dei valori dei pg e i pg stessi
    """
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()

    with open(file_path, 'r') as json_file:
        data = json.load(json_file)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Populate the first and second columns with data from JSON
    riga = 0
    for descrizione, valore in data.items():
        riga += 1
        sheet.cell(row=riga, column=1).value = descrizione
        sheet.cell(row=riga, column=2).value = valore

    # Save the workbook
    nome = file_path.split("\\")[-1].split(".")[0] + ".xlsx"
    workbook.save(nome)

def excel_to_json_pg():
    import json
    from openpyxl import load_workbook

    root = Tk()
    root.withdraw()
    file_path = askopenfilename()

    # Load the workbook
    workbook = load_workbook(file_path)

    # Select the active sheet
    sheet = workbook.active

    # Create an empty dictionary to store the data
    data = {}

    # Iterate over rows in the first column
    for row in sheet.iter_rows(values_only=True):
        index = row[0]
        value = row[1]

        # Add the index and value to the dictionary
        data[str(index)] = value

    # Save the data as JSON
    nome = file_path.split("\\")[-1].split(".")[0] + ".json"
    with open(nome, 'w') as json_file:
        json.dump(data, json_file, indent=4)

def json_to_excel_item():
    # va bene anche per le skill
    # Load JSON data from file
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()

    with open(file_path, 'r') as json_file:
        data = json.load(json_file)

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Populate the first and second columns with data from JSON
    row = 1
    for obj, values in data.items():
        for col, value in enumerate(values,1):
            sheet.cell(row=row, column=col).value = value
        row += 1

    # Save the workbook
    nome = file_path.split("\\")[-1].split(".")[0] + ".xlsx"
    workbook.save(nome)

def excel_to_json_item():
    # va bene anche per le skill
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active sheet
    sheet = workbook.active

    # Create an empty dictionary to store the data
    data = {}

    # Iterate over rows
    riga = 0
    for row in sheet.iter_rows(values_only=True):
        riga += 1
        obj = str(riga)
        values = row[0:]

           # Add the object and values to the dictionary
        data[obj] = values

    # Save the data as JSON
    nome = file_path.split("\\")[-1].split(".")[0] + ".json"
    with open(nome, 'w') as json_file:
        json.dump(data, json_file, indent=4)

def excel_to_json_npc():
    import json
    from openpyxl import load_workbook

    root = Tk()
    root.withdraw()
    file_path = askopenfilename()

    # Load the workbook
    workbook = load_workbook(file_path)

    # Select the active sheet
    sheet = workbook.active

    # Create an empty dictionary to store the data
    data = {}

    # Iterate over rows in the first column
    indice = 0
    for colonna in range(sheet.max_column):
        indice += 1
        npc = {}
        for riga in sheet.iter_rows(values_only=True):
            index = riga[0]
            try:
                value = riga[indice]
            except:
                break
            # Add the index and value to the dictionary
            npc[str(index)] = value
        try:
            data[str(npc["nome_valore_excel"])] = npc
        except:
            break
    # Save the data as JSON
    nome = file_path.split("\\")[-1].split(".")[0] + ".json"
    with open(nome, 'w') as json_file:
        json.dump(data, json_file, indent=4)

def json_to_excel_npc():
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()

    with open(file_path, 'r') as json_file:
        data = json.load(json_file)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Populate the first and second columns with data from JSON
    indice = 1
    for npc in data:
        indice += 1
        riga = 0
        for descrizione, valore in data[npc].items():
            riga += 1
            sheet.cell(row=riga, column=1).value = descrizione
            sheet.cell(row=riga, column=indice).value = valore

    # Save the workbook
    nome = file_path.split("\\")[-1].split(".")[0] + ".xlsx"
    workbook.save(nome)

def crea_negozi():
    from Moduli.Logica import Oggetti
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()

    with open(file_path, 'r', encoding='utf-8') as json_file:
        data_totale = json.load(json_file)

    for zona in data_totale:
        for negozio in data_totale[zona]:
            oggetti_negozio = Oggetti.estrai_negozio(tipo_negozio=negozio["tipo"].lower(), livello=negozio["livello"])
            negozio["oggetti"] = oggetti_negozio

    nome = file_path.split("\\")[-1].split("_e.")[0] + ".json"
    with open(nome, 'w') as json_file:
        json.dump(data_totale, json_file, indent=4)

def extra1(): #in questo caso, cambia dati dell'xlsx di oggetti
    import openpyxl
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()
    # Load your workbook and select the worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming you're working with the first sheet
    json_tipi = {"martello": ["corta", "potente", "Colpo potente", ""],
                 "tirapugni": ["corta", "media", "no slot/hidden", ""],
                 "nunchaku": ["corta", "precisa", "1 reroll/turno", ""],
                 "coltello": ["corta", "media", "tira arma", ""],
                 "daga": ["corta", "precisa", "dmg cont./perf Opzionale", ""],
                 "armblade": ["corta", "potente", "+2 def", "attore_attivo['difesa_item']+=2"],
                 "stiletto": ["corta", "potente", "Colpo furtivo", ""],
                 "shiv": ["corta", "precisa", "no slot/hidden", ""],
                 "kriss": ["corta", "media", "+1atk / dmg tagl. Opz.", "attore_attivo['attacco_item'] +=1"],
                 "mazza": ["media", "media", "Sbilancia", ""],
                 "mazzafrusta": ["media", "precisa", "Ruota", ""],
                 "kusarigama": ["media", "precisa", "lancio peso", ""],
                 "spadalunga": ["media", "precisa", "dmg cont./perf Opzionale", ""],
                 "sciabola": ["media", "potente", "Parata", ""],
                 "katana": ["media", "media", "Sanguina", ""],
                 "fioretto": ["media", "media", "Affondo", ""],
                 "estoc": ["media", "potente", "Colpo abile", ""],
                 "bastone": ["lunga", "precisa", "Parata", ""],
                 "martellodaguerra": ["lunga", "potente", "Sbilancia", ""],
                 "bastoneconpesi": ["lunga", "media", "Colpo dal basso", ""],
                 "asciaaduemani": ["lunga", "media", "+1 atk / dmg cont. Opzionale",
                                   "attore_attivo['attacco_item'] +=1"],
                 "spadone": ["lunga", "precisa", "dmg cont./perf Opzionale", ""],
                 "zweihander": ["lunga", "potente", "Ruota", ""],
                 "lancia": ["lunga", "media", "Sanguina", ""],
                 "picca": ["lunga", "potente", "+1atk / dmg tagl. Opz.", "attore_attivo['attacco_item'] +=1"],
                 "beccodicorvo": ["lunga", "precisa", "+1atk / dmg cont. Opz.", "attore_attivo['attacco_item'] +=1"],
                 "coltellodalancio": ["corta", "medie", "Tiro Rapido", ""],
                 "accettadalancio": ["corta", "potente", "Tiro Attento", ""],
                 "shuriken": ["corta", "precisa", "Counterspell", ""],
                 "balestra": ["media", "precisa", "Focus", ""],
                 "balestraaripetizione": ["media", "media", "Doppio Missile", ""],
                 "arcocorto": ["media", "media", "Tiro Rapido", ""],
                 "arcolungo": ["lunga", "potente", "+2 atk", "attore_attivo['attacco_item'] +=2"],
                 "arcocomposito": ["lunga", "precisa", "Tiro Attento", ""],
                 "chukonu": ["lunga", "potente", "Triplo Missile", ""],
                 "tonfa": ["corta", "precisa", "+2 def", "attore_attivo['difesa_item']+=2"],
                 "tridente": ["lunga", "potente", "Tira arma", ""],
                 "accetta": ["corta", "potente", "tira arma", ""],
                 "ascia": ["media", "potente", "", ""],
                 "maninude": ["maninude", "maninude", "", ""],
                 "natura1": ["corta", "precisa", "", ""],
                 "natura2": ["media", "medie", "", ""],
                 "natura3": ["lunga", "potente", "", ""],
                 "bastonemagico": ["lunga", "precisa", "", ""]}
    # Loop through the rows and update the cells

    for row in ws.iter_rows(min_row=2, max_row=10000, min_col=1, max_col=ws.max_column):
        # Assuming "EFFETTO1" is in column J, "EFFETTO2" is in column K, and "EFFETTO3" is in column L
        effetto1_cell = row[10]
        effetto2_cell = row[11]
        effetto3_cell = row[12]
        pa_atk_cell = row[6]
        if row[2].value in json_tipi.keys():
        # Update the cells
            effetto1_cell.value = f"Personaggio.attacco + {effetto1_cell.value}" if type(effetto1_cell.value) is int or "-" not in effetto1_cell.value else f"Personaggio.attacco - {effetto1_cell.value.replace('-','')}"
            effetto2_cell.value = f"Personaggio.tier_danno + {effetto2_cell.value}" if type(effetto2_cell.value) is int or "-" not in effetto2_cell.value else f"Personaggio.tier_danno - {effetto2_cell.value.replace('-','')}"
            pa_atk_cell.value = f"{effetto3_cell.value}"
            try:
                int(effetto3_cell.value)
                effetto3_cell.value = ''
            except:
                pass
        elif row[2].value in ["armatura","chainmail","veste","scudo"]:
            dif = effetto1_cell.value if effetto1_cell.value not in ["Vuoto", "", None] else "0"
            pa = effetto2_cell.value if effetto2_cell.value not in ["Vuoto", "", None] else "0"
            effetto1_cell.value = f"Personaggio.difesa + {dif}" if type(dif) is int or "-" not in dif else f"Personaggio.difesa - {dif.replace('-','')}"
            effetto2_cell.value = f"Personaggio.pa - {pa}" if type(pa) is int or "-" not in pa else f"Personaggio.pa - {pa.replace('-','')}"

        for i in range(35):
            try:
                ell = row[i]
                val = ell.value


            except IndexError:
                pass
    wb.save("updated_file.xlsx")
def extra2():
    import openpyxl
    root = Tk()
    root.withdraw()
    file_path = askopenfilename()
    # Load your workbook and select the worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming you're working with the first sheet
    diz = {}
    for row in ws.iter_rows(min_row=2, max_row=10000, min_col=1, max_col=ws.max_column):
        try:
            int(row[0].value)
            nuovo_dict = {"id" : row[0].value,
                        "nome" : row[1].value,
                        "tipo_1" : row[2].value,
                        "tipo_2" : row[3].value,
                        "tipo_3" : row[4].value,
                        "tipo_4" : row[5].value,
                        "pa_per_attacco" : row[6].value,
                        "descrizione" : row[7].value,
                        "valore" : row[8].value,
                        "peso" : row[9].value,
                        "effetto_1" : row[10].value,
                        "effetto_2" : row[11].value,
                        "effetto_3" : row[12].value,
                        "effetto_4" : row[13].value,
                        "effetto_5" : row[14].value,
                        "effetto_6" : row[15].value,
                        "effetto_7" : row[16].value,
                        "effetto_8" : row[17].value,
                        "effetto_9" : row[18].value,
                        "effetto_10" : row[19].value,
                        "lv_loot" : row[20].value,
                        "rarita" : row[21].value
                          }
            diz[str(row[0].value)] = nuovo_dict
        except:
            pass
    nome = file_path.split("\\")[-1].split(".")[0] + ".json"
    with open(nome, 'w') as json_file:
        json.dump(diz, json_file, indent=4)


window = Tk()

# Create the buttons
button1 = Button(window, text="json_to_excel_item_skill", command=json_to_excel_item)
button2 = Button(window, text="excel_to_json_item_skill", command=excel_to_json_item)
button3 = Button(window, text="excel_to_json_npc", command=excel_to_json_npc)
button4 = Button(window, text="json_to_excel_npc", command=json_to_excel_npc)
button5 = Button(window, text="excel_to_json_pg", command=excel_to_json_pg)
button6 = Button(window, text="json_to_excel_pg", command=json_to_excel_pg)
button7 = Button(window, text="crea_negozi", command=crea_negozi)
button8 = Button(window, text="EXTRA1", command=extra1)
button9 = Button(window, text="EXTRA2", command=extra2)


button1.pack()
button2.pack()
button3.pack()
button4.pack()
button5.pack()
button6.pack()
button7.pack()
button8.pack()
button9.pack()

window.mainloop()

