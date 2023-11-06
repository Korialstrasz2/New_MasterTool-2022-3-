import ast
import json
import openpyxl
import os
import threading
from datetime import datetime
from time import sleep

import Moduli.SharedData as Shared
from Moduli.Grafica import Popups_ScreenPGImportato, Popups_Zaino
from Moduli.Logica import EquipAttore, Inizializza, ImportazioneAttori, GestioneSkill, Oggetti

Shared.timer_start = datetime.now()
timeout_save = False


def crea_excel_sunto_pg(*args):
    print(Shared.pg_selezionato)
    cartella = f"{Shared.path_cartella_shared}/Excel_PG"
    if not os.path.exists(cartella):
        os.makedirs(cartella)


    data = Shared.pg_selezionato

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Populate the first and second columns with data from JSON
    riga = 0
    caratteristiche = ["forza_base","resistenza_base","velocita_base","agilita_base","intelligenza_base",
                       "concentrazione_base","personalita_base","saggezza_base","fortuna_base"]
    for descrizione in Shared.dati_base_pg:
        valore = data[descrizione]
        if descrizione in caratteristiche:
            descrizione = descrizione.replace("_base","_tot")
            valore = data[descrizione]
        if descrizione.startswith("equip") or descrizione.startswith("id_") or descrizione.startswith("zaino_"):
            if valore.startswith("id:"):
                valore = valore[3:]
            try:
                int(valore)
                valore = Oggetti.trova_oggetto(valore)["NOME"]
            except:
                pass
        if str(valore) not in ["","None","0", "Vuoto"] and not descrizione.startswith("formul"):
            riga += 1
            sheet.cell(row=riga, column=1).value = descrizione
            sheet.cell(row=riga, column=2).value = valore


    dataskill = Shared.skill_importate

    row = 1
    for obj, values in dataskill.items():
        colonna = 4
        for value in values:
            value = values[value]
            sheet.cell(row=row, column=colonna).value = value
            colonna += 1
            if colonna == 7:
                if row == 1:
                    sheet.cell(row=row, column=colonna).value = "PE SPESI"
                for id_speso in Shared.skill_sbloccate_pg:
                    if str(id_speso) == str(obj):
                        sheet.cell(row=row, column=colonna).value = "spesi: " + str(Shared.skill_sbloccate_pg[id_speso])
                colonna += 1
        row += 1
    # Save the workbook
    nome = cartella + "/" + data["nomepg"] + ".xlsx"
    workbook.save(nome)



def check_cambiamenti_attori_invio_rapido(nomepg):
    EquipAttore.equip_npc(nomepg)
    attorenpc = Shared.pg_png_inizializzati[nomepg]
    numero_npc = 0
    da_passare = {}
    for npc_attivo in range(1, 11):
        nome = Shared.npc_numeri_assegnati[npc_attivo]['nome_in_uso']
        if nomepg == nome:
            numero_npc = npc_attivo

    da_passare['numero_npc'] = numero_npc
    da_passare['nome_in_uso'] = attorenpc['nome_in_uso']
    for chiave in Shared.dati_base_pg:
        da_passare[chiave] = attorenpc[chiave]
    return da_passare

def scrivi_in_cartella_condivisa(nomepg):
    dati = check_cambiamenti_attori_invio_rapido(nomepg)
    file_name = datetime.now()
    file_name = str(file_name)
    file_name = file_name.replace(":", "-")
    file_name = file_name.replace(" ", "_")
    file_name = file_name.replace(".", "-")

    nomepg = dati["nome_in_uso"]
    cartelladump = f"{Shared.path_cartella_shared}/{nomepg}"

    if not os.path.exists(cartelladump):
        os.makedirs(cartelladump)
    files = os.listdir(cartelladump)
    files_with_time = []
    for file in files:
        file_path = os.path.join(cartelladump, file)
        file_time = os.path.getmtime(file_path)
        files_with_time.append((file_time, file))
    files_with_time.sort(reverse=True)
    files = [file for _, file in files_with_time if not file.startswith("__")]
    if len(files) > 7:
        os.remove(os.path.join(cartelladump, str(files[-1])))

    nomecompletofile = f"{nomepg}-{file_name}.json"
    with open(cartelladump + "/" + nomecompletofile, "w") as json_file:
        json.dump(dati, json_file, indent=4)
    if nomepg not in Shared.file_letti_per_pg:
        Shared.file_letti_per_pg[nomepg] = []
    Shared.file_letti_per_pg[nomepg].append(nomecompletofile)


def salva_pgnpc(nome, share=True):
    for inizializzato in Shared.pg_png_inizializzati:
        if inizializzato == nome:
            newdic = {}
            for dato in Shared.dati_base_pg:
                newdic[dato] = Shared.pg_png_inizializzati[inizializzato][dato]
            file_locale = Shared.path_PG_e_Unici + "\\" + Shared.pg_png_inizializzati[inizializzato][
                "nome_in_uso"] + ".json"
            with open(file_locale, "w") as json_file:
                json.dump(newdic, json_file, indent=4)
            if share:
                scrivi_in_cartella_condivisa(nome)
            break

def crea_bottoni_base():
    dati = {"bottone_1": {"nome":"", "nome_effetto": "", "descrizione_effetto":"","codice_effetto":"","titolo_popup":"",
                          "contenuto_popup":"","link":"","campo_PF":"","campo_Mana":"","campo_Energia":"","campo_Potere":"",
                          "campo_Punti_Azione":"","campo_Bonus_Atk":"","campo_Bonus_Tier":"","campo_Bonus_DMG":""},
            "bottone_2": {"nome":"", "nome_effetto": "", "descrizione_effetto":"","codice_effetto":"","titolo_popup":"",
                          "contenuto_popup":"","link":"","campo_PF":"","campo_Mana":"","campo_Energia":"","campo_Potere":"",
                          "campo_Punti_Azione":"","campo_Bonus_Atk":"","campo_Bonus_Tier":"","campo_Bonus_DMG":""},
            "bottone_3": {"nome":"", "nome_effetto": "", "descrizione_effetto":"","codice_effetto":"","titolo_popup":"",
                          "contenuto_popup":"","link":"","campo_PF":"","campo_Mana":"","campo_Energia":"","campo_Potere":"",
                          "campo_Punti_Azione":"","campo_Bonus_Atk":"","campo_Bonus_Tier":"","campo_Bonus_DMG":""},
            "bottone_4": {"nome":"", "nome_effetto": "", "descrizione_effetto":"","codice_effetto":"","titolo_popup":"",
                          "contenuto_popup":"","link":"","campo_PF":"","campo_Mana":"","campo_Energia":"","campo_Potere":"",
                          "campo_Punti_Azione":"","campo_Bonus_Atk":"","campo_Bonus_Tier":"","campo_Bonus_DMG":""},
            "bottone_5": {"nome":"", "nome_effetto": "", "descrizione_effetto":"","codice_effetto":"","titolo_popup":"",
                          "contenuto_popup":"","link":"","campo_PF":"","campo_Mana":"","campo_Energia":"","campo_Potere":"",
                          "campo_Punti_Azione":"","campo_Bonus_Atk":"","campo_Bonus_Tier":"","campo_Bonus_DMG":""},
            "bottone_6": {"nome":"", "nome_effetto": "", "descrizione_effetto":"","codice_effetto":"","titolo_popup":"",
                          "contenuto_popup":"","link":"","campo_PF":"","campo_Mana":"","campo_Energia":"","campo_Potere":"",
                          "campo_Punti_Azione":"","campo_Bonus_Atk":"","campo_Bonus_Tier":"","campo_Bonus_DMG":""}}
    nome_pg = Shared.pg_selezionato["nome_in_uso"]
    cartelladump = f"{Shared.path_cartella_shared}/{nome_pg}/bottoni"
    if not os.path.exists(cartelladump):
        os.makedirs(cartelladump)
    cartelladump2 = f"{Shared.path_PG_e_Unici}/bottoni/"
    if not os.path.exists(cartelladump2):
        os.makedirs(cartelladump2)

    pathcompleto = Shared.path_cartella_shared + f"/{nome_pg}/bottoni/{nome_pg}bottoni.json"
    with open(pathcompleto, "w") as json_file:
        json.dump(dati, json_file, indent=4)
    pathcompleto = Shared.path_PG_e_Unici + f"/bottoni/{nome_pg}bottoni.json"
    with open(pathcompleto, "w") as json_file:
        json.dump(dati, json_file, indent=4)
    return dati

def salva_skill_sbloccate():
    dati = Shared.skill_sbloccate_pg
    file_locale = Shared.path_PG_e_Unici + "\\" + Shared.pg_selezionato["nome_in_uso"] + "skill" + ".json"
    with open(file_locale, "w") as json_file:
        json.dump(dati, json_file, indent=4)


    file_name = datetime.now()
    file_name = str(file_name)
    file_name = file_name.replace(":", "-")
    file_name = file_name.replace(" ", "_")
    file_name = file_name.replace(".", "-")

    nomepg = Shared.pg_selezionato["nome_in_uso"]
    cartelladump = f"{Shared.path_cartella_shared}/{nomepg}/skill"

    if not os.path.exists(cartelladump):
        os.makedirs(cartelladump)
    files = os.listdir(cartelladump)
    files_with_time = []
    for file in files:
        file_path = os.path.join(cartelladump, file)
        file_time = os.path.getmtime(file_path)
        files_with_time.append((file_time, file))
    files_with_time.sort(reverse=True)
    files = [file for _, file in files_with_time]
    if len(files) > 7:
        os.remove(os.path.join(cartelladump, str(files[-1])))

    nomecompletofile = f"{nomepg}skill-{file_name}.json"
    nomecompletofile2 = f"{nomepg}skill.json"
    with open(cartelladump + "/" + nomecompletofile, "w") as json_file:
        json.dump(dati, json_file, indent=4)
    with open(cartelladump + "/" + nomecompletofile2, "w") as json_file:
        json.dump(dati, json_file, indent=4)

def salva_tuttobup5():
    for pg in Shared.pg_png_inizializzati:
        salva_pgnpc(pg, share=False)
    print("back up 5 min")
    t5b = threading.Timer(150.0, salva_tuttobup5)
    t5b.daemon = True
    t5b.start()


t5 = threading.Timer(150.0, salva_tuttobup5)
t5.daemon = True
t5.start()


def timer_salva():
    global timeout_save
    timeout_save = True
    sleep(5)
    timeout_save = False


def aggiungi_pg_da_json(*args):
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    file_selected = filedialog.askopenfilename()
    if len(str(file_selected)) > 0:
        with open(file_selected, 'r') as json_file:
            data = json.load(json_file)
            ImportazioneAttori.inizializza_pg_esterno(data)
            Popups_ScreenPGImportato.popup_check_differenze_pgnpc(data)


def flag_esporta_npc(*args):
    flagfilepresente = False
    for file in os.listdir(Shared.path_cartella_shared):
        if file == "flagexportnpc.txt":
            flagfilepresente = True
            pathfile = os.path.join(Shared.path_cartella_shared,file)
            os.remove(pathfile)
    if flagfilepresente == False:
        with open(Shared.path_cartella_shared + f"/flagexportnpc.txt", "w") as f:
            f.write(str(Shared.npc_numeri_assegnati))



def importa_tutti_npc(*args):
    import ast
    pathfile = os.path.join(Shared.path_cartella_shared, "flagexportnpc.txt")
    if "flagexportnpc.txt" in os.listdir(Shared.path_cartella_shared):
        with open(pathfile, "r") as f:
            dictinput = ast.literal_eval(f.read())
        for npc in dictinput:
            nome_in_uso = dictinput[npc]["nome_in_uso"]
            Shared.pg_png_inizializzati[nome_in_uso] = dictinput[npc]
            EquipAttore.equip_npc(nome_in_uso)
            ImportazioneAttori.assegna_npc(nome_in_uso, npc)


def flag_esporta_db_oggetti(*args):
    flagfilepresente = False
    for file in os.listdir(Shared.path_cartella_shared):
        if file == "flagexportdboggetti.json":
            flagfilepresente = True
            pathfile = os.path.join(Shared.path_cartella_shared,file)
            os.remove(pathfile)
    if flagfilepresente == False:
        with open(Shared.path_cartella_shared + f"/flagexportdboggetti.json", "w") as json_file:
            json.dump(Shared.file_oggetti, json_file, indent=4)


def flag_importa_db_oggetti(*args):
    Oggetti.importa_db_oggetti()

def flag_esporta_db_skill(*args):
    flagfilepresente = False
    for file in os.listdir(Shared.path_cartella_shared):
        if file == "flagexportdbskill.json":
            flagfilepresente = True
            pathfile = os.path.join(Shared.path_cartella_shared,file)
            os.remove(pathfile)
    if flagfilepresente == False:
        with open(Shared.path_cartella_shared + f"/flagexportdbskill.json", "w") as json_file:
            json.dump(Shared.skill_importate, json_file, indent=4)


def flag_importa_db_skill(*args):
    GestioneSkill.esporta_db_skill()


def salvazaino(*args):
    Zaino = Popups_Zaino.Zaino
    for slot_m in range(0, 20):
        if eval(f"Zaino.boxidslot{slot_m + 1}m.text") != "":
            exec(f"Shared.pg_selezionato['zaino_slot_{slot_m + 1}'] = 'id:'+Zaino.boxidslot{slot_m + 1}m.text")
        else:
            exec(f"Shared.pg_selezionato['zaino_slot_{slot_m + 1}'] = Zaino.boxnomeslot{slot_m + 1}m.text")

    for slot_n in range(0, 20):
        if eval(f"Zaino.boxidslot{slot_n + 1}n.text") != "":
            exec(f"Shared.pg_selezionato['zaino_slot_{slot_n + 21}'] = 'id:'+Zaino.boxidslot{slot_n + 1}n.text")
        else:
            exec(f"Shared.pg_selezionato['zaino_slot_{slot_n + 21}'] = Zaino.boxnomeslot{slot_n + 1}n.text")

    Shared.pg_selezionato['id_arma_1'] = Zaino.boxidslotar1.text
    Shared.pg_selezionato['id_arma_2'] = Zaino.boxidslotar2.text
    Shared.pg_selezionato['id_armatura'] = Zaino.boxidslotse1.text
    Shared.pg_selezionato['id_scudo'] = Zaino.boxidslotse2.text
    Shared.pg_selezionato['id_chainmail'] = Zaino.boxidslotse3.text
    Shared.pg_selezionato['id_veste'] = Zaino.boxidslotse4.text
    Shared.pg_selezionato['monete'] = Zaino.boxmonete.text
    for note in range(1, 5):
        exec(f"Shared.pg_selezionato['note_zaino_{note}'] = Zaino.boxnote{note}.text")
    salva_pgnpc(Shared.pg_selezionato["nome_in_uso"])

